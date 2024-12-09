import openpyxl


class ExcelReader:
    def __init__(self, file_path):
        self.file_path = file_path

    def get_data(self, sheet_name):
        workbook = openpyxl.load_workbook(self.file_path)
        sheet = workbook[sheet_name]
        data = []
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
            data.append(row)
        return data